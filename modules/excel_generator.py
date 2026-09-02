import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

REPORT_FOLDER = Path("data/reports")
REPORT_FOLDER.mkdir(parents=True, exist_ok=True)

HEADER_FILL = PatternFill("solid", fgColor="1E40AF")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def write_dataframe(ws, df):
    """Write any dataframe safely to an Excel worksheet."""

    for col_num, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_num, value=str(col_name))
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for row in df.fillna("").itertuples(index=False):
        ws.append(list(row))

    for column_cells in ws.columns:
        length = max(len(str(c.value)) if c.value is not None else 0 for c in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(length + 4, 35)


def create_workforce_excel_report(
    df,
    dataset_name="Workforce",
    quality=None,
    ai_summary=""
):
    """
    Generates enterprise Excel workbook.
    Returns generated file path.
    """

    wb = Workbook()


    dashboard = wb.active
    if dashboard is not None:
        dashboard.title = "Executive Dashboard"

        dashboard["A1"] = "InsightForge AI Workforce Executive Report"
        dashboard["A1"].font = Font(size=16, bold=True)

        quality = quality or {}

        metrics = [
            ("Dataset", dataset_name),
            ("Employees", len(df)),
            ("Departments", df["Department"].fillna("Unknown").nunique() if "Department" in df else 0),
            ("Cities", df["City"].fillna("Unknown").nunique() if "City" in df else 0),
            ("Average Salary", round(pd.to_numeric(df["Salary"], errors="coerce").fillna(0).mean(), 2) if "Salary" in df else 0),
            ("Attendance %", round(pd.to_numeric(df["Attendance_Percentage"], errors="coerce").fillna(0).mean(), 2) if "Attendance_Percentage" in df else 0),
            ("Performance Rating", round(pd.to_numeric(df["Performance_Rating"], errors="coerce").fillna(0).mean(), 2) if "Performance_Rating" in df else 0),
            ("Attrition Rate %", round((df["Attrition_Status"].fillna("") == "Left").mean() * 100, 2) if "Attrition_Status" in df else 0),
            ("Quality Score", quality.get("score", 100))
        ]

        row = 3
        for label, value in metrics:
            dashboard[f"A{row}"] = label
            dashboard[f"B{row}"] = value
            row += 1


    employee_sheet = wb.create_sheet("Employee Master")
    write_dataframe(employee_sheet, df)


    if "Department" in df.columns:

        dept_sheet = wb.create_sheet("Department Analysis")

        dept = (
            df.groupby("Department")
            .agg(
                Employees=("Employee_ID", "count"),
                Average_Salary=("Salary", "mean"),
                Average_Performance=("Performance_Rating", "mean"),
                Average_Attendance=("Attendance_Percentage", "mean")
            )
            .reset_index()
            .fillna(0)
        )

        write_dataframe(dept_sheet, dept)


    if "Salary" in df.columns:

        salary_sheet = wb.create_sheet("Salary Analysis")

        salary_df = df[["Employee_Name", "Department", "Job_Role", "Salary"]].copy()
        salary_df["Salary"] = pd.to_numeric(salary_df["Salary"], errors="coerce").fillna(0)

        write_dataframe(salary_sheet, salary_df.sort_values("Salary", ascending=False))

  

    if "Attendance_Percentage" in df.columns:

        attendance_sheet = wb.create_sheet("Attendance Analysis")

        attendance_df = df[
            ["Employee_Name", "Department", "Attendance_Percentage"]
        ].copy()

        attendance_df["Attendance_Percentage"] = pd.to_numeric(
            attendance_df["Attendance_Percentage"],
            errors="coerce"
        ).fillna(0)

        write_dataframe(attendance_sheet, attendance_df)


    if "Performance_Rating" in df.columns:

        performance_sheet = wb.create_sheet("Performance Analysis")

        performance_df = df[
            [
                "Employee_Name",
                "Department",
                "Performance_Rating",
                "Training_Hours",
                "Promotion_Status"
            ]
        ].copy()

        performance_df = performance_df.fillna(0)

        write_dataframe(performance_sheet, performance_df)


    if "Attrition_Status" in df.columns:

        attrition_sheet = wb.create_sheet("Attrition Analysis")

        attrition_df = df[
            [
                "Employee_Name",
                "Department",
                "City",
                "Salary",
                "Experience_Years",
                "Attrition_Status"
            ]
        ].copy()

        write_dataframe(attrition_sheet, attrition_df.fillna("Unknown"))


    ai_sheet = wb.create_sheet("AI Workforce Summary")

    ai_sheet["A1"] = "AI Executive Workforce Summary"
    ai_sheet["A1"].font = Font(size=14, bold=True)

    if ai_summary:
        lines = str(ai_summary).split("\n")
        for i, line in enumerate(lines, start=3):
            ai_sheet.cell(row=i, column=1).value = line
    else:
        ai_sheet["A3"] = "No AI Summary Generated."


    output_file = REPORT_FOLDER / f"{dataset_name}_Executive_Report.xlsx"
    wb.save(output_file)

    return str(output_file)
